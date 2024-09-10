import openpyxl

SAVE_LOCATION = 'C:/Users/doyle/OneDrive - University of Toronto/RSS/Automations/Duty Solver/Solved/'

path_original = ''
path_final = ''
origin_row = 0
origin_column = 0

end_row = 0
end_column = 0

skipped_days = []

first_pass = False
error_messages = []


def get_dimensions(ws):
    # gets the dimensions of the worksheet!
    #
    global origin_row
    global origin_column
    global end_row
    global end_column


    # First to find the top left corner!
    found_points = False
    for row in range(1, ws.max_row):
        for col in range(0, ws.max_column):
            if not found_points:
                entry = ws[row][col].value
                if entry:
                    try:
                        entry = entry.strip().lower()
                        if entry == "points":
                            found_points = True
                            origin_row = row + 1
                            origin_column = col + 1
                    except:
                        pass
    if not found_points:
        print(f'Unable to find cell containing entry "points" ')
        error_messages.append(f'')
        error_messages.append(f'* ERROR: Unable to find cell containing word "points"')
        return False

    # print(f'think start is {origin_row} {origin_column}')


    # Now to find the bottom right corner!
    # First, find the cell containing the word INSTRUCTIONS
    found_instructions = False
    for row in range(1, ws.max_row):
        # now we're going through the entire first column!
        # We have to stop once we find the line containing
        # the phrase "Instructions"
        if not found_instructions:
            entry = ws[row][0].value
            if entry.strip() == "Instructions":
                found_instructions = True
                end_row = row - 1

    if not found_instructions:
        print(f'Unable to locate cell containing word "Instructions"')

        error_messages.append(f'')
        error_messages.append(f'* ERROR: Unable to find cell containing word "Instructions"')

        return False

    # Now we find the cell where the date is empty!
    found_end_date = False
    for col in range(0, ws.max_column):
        if not found_end_date:
            entry = ws[2][col].value
            if not entry:
                found_end_date = True
                end_column = col - 1

    if not found_end_date:
        print(f'Unable to locate last column')
        print(f'inferring last column as final column')

        error_messages.append(f'')
        error_messages.append(f'* ERROR: Unable to locate the final column!')
        error_messages.append(f'* Inferring last column as final column')

        end_column = ws.max_column - 1





def get_origin(ws):
    # Returns the origin for the sheet!
    # 3, 2 would mean row 3 column 2, or cell C3
    # (column indexes by 0, row doesn't)
    if origin_row == 0 and origin_column == 0:
        if get_dimensions(ws) == False:
            return False

    return origin_row, origin_column

def get_end(ws):
    # Takes the worksheet and finds the bottom right corner
    # of the data!
    #
    # 4, 6 would mean row 4 column 6, or cell G4
    # AL 10
    global end_row
    global end_column

    if end_row == 0 and end_column == 0:
        get_dimensions(ws)
    # print(f'think end is {end_row} {end_column}')
    return end_row, end_column


def process_entry(entry):
    val = 0
    if entry:
        entry_array = entry.split()
        try:
            val = int(entry_array[-1])
        except:
            print(f'ERROR ERROR ERROR ERROR ERROR')
            print(f'invalid entry!')
            print(f'entry:{entry}.')
            print(f'entry array:{entry_array}')
            print(f'ERROR ERROR ERROR ERROR ERROR')

            error_messages.append(f'* ERROR: Invalid Entry!')
            error_messages.append(f'entry:{entry}.')

    return val


def open_notebook(path):
    # Takes a path to a .xlsx file.
    # Returns tuple of availability, dates, and names.
    reset_variables()
    global path_original
    global first_pass
    path_original = path
    if not first_pass:
        global run_data
        run_data_directory = SAVE_LOCATION + "run_data.txt"
        run_data = open(run_data_directory, "a+")

        first_pass = True


    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(path)

    # Define variable to read sheet
    dataframe1 = dataframe.active


    # Start by getting all the dates
    availability = []
    availability1 = []
    names = []

    # create data structure for availability!
    # Simultaneously add names to name list
    found_instructions = False
    for row in range(2, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, 1):
            # now we're going through the entire first column!
            # We have to stop once we find the line containing
            # the phrase "Instructions"
            if not found_instructions:
                entry = col[row].value.strip()
                if entry.strip() == "Instructions":
                    found_instructions = True
                else:
                    names.append(entry)
                    availability.append([])
                    availability1.append([])

    if get_origin(dataframe1) == False:
        return False, False
    r, c = get_origin(dataframe1)
    r_f, c_f = get_end(dataframe1)

    global skipped_days
    for col in range(c, c_f + 1):
        empty_row = True
        for row in range(r, r_f + 1):
            val = process_entry(dataframe1[row][col].value)
            availability1[row - r].append(val)
            if val != 0:
                empty_row = False
        if empty_row:
            print(f'found empty column')
            error_messages.append(f'Warning: Found empty column. Double check solution is valid.')

            skipped_days.append(col)
            for row in range(r, r_f + 1):
                availability1[row - r].pop()

    print(f'skipping: {skipped_days}')
    # Saves the dataframe so we can re-open it later!
    dataframe.save(path)

    return availability1, names


def save_workbook(availability):
    # Saves the workbook based on the final_availability!
    #
    # Makes a copy of the workbook into Solved_Workbooks.
    # Renames the sheet name (not workbook name)

    global path_original
    global path_final
    global skipped_days
    path_final = SAVE_LOCATION
    original_name = path_original.split('/')[-1]
    path_final += original_name[:-5] + "_SOLVED" + original_name[-5:]

    wb = openpyxl.load_workbook(path_original)
    sheet = wb.active
    wb.save(path_original)

    r, c = get_origin(sheet)

    # s helps us track the days we skip!
    s = 0
    for col in range(c, min(len(availability[0]) + c, sheet.max_column)):
        # This helps us skip days
        if col + s in skipped_days:
            s += 1

        for row in range(r, min(len(availability) + r, sheet.max_row)):
            if availability[row-r][col-c] == 0:
                sheet[row][col+s].value = None


    original_title = sheet.title.split()
    final_title = ''
    for i in range(len(original_title)):
        if i != 2:
            final_title += original_title[i]
        else:
            final_title += 'SOLVED'
        final_title += ' '
    final_title = final_title[:-1]
    sheet.title = final_title

    wb.save(path_final)






def reset_variables():
    # resets all global variables
    global path_original
    global path_final
    global origin_row
    global origin_column
    global end_row
    global end_column
    global skipped_days

    path_original = ''
    path_final = ''
    origin_row = 0
    origin_column = 0

    end_row = 0
    end_column = 0

    skipped_days = []





if __name__ == "__main__":
    availability, dates, names = open_notebook("Excel_Files/Example_Workbooks/Test.xlsx")

    print(f'-----------------------')
    print(dates)
    print(names)
    print(f'---')
    for row in availability:
        print(row)

    print(len(availability[0]))
