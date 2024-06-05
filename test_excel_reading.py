import openpyxl



def process_entry(entry):
    val = 0
    if entry:
        entry_array = entry.split()
        try:
            val = int(entry_array[-1])
        except:
            print(f'ERROR ERROR ERROR ERROR ERROR')
            print(f'invalid entry!')
            print(f'entry:{entry}.')
            print(f'entry array:{entry_array}')
            print(f'ERROR ERROR ERROR ERROR ERROR')
    return val


def process_notebook(path):
    # Takes a path to a .xlsx file.
    # Returns tuple of availability, dates, and names.


    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(path)

    # Define variable to read sheet
    dataframe1 = dataframe.active


    # Start by getting all the dates
    dates = []
    availability = []
    names = []

    # create data structure for availability!
    # Simultaneously add names to name list
    found_instructions = False
    for row in range(2, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, 1):
            # now we're going through the entire first column!
            # We have to stop once we find the line containing
            # the phrase "Instructions"
            if not found_instructions:
                entry = col[row].value.strip()
                if entry.strip() == "Instructions":
                    found_instructions = True
                else:
                    names.append(entry)
                    availability.append([])

    for col in dataframe1.iter_cols(3, dataframe1.max_column):
        # the if col[1].value essentially cuts off once we find the last row.
        # It only proceeds if the date row is not None!
        if col[1].value:
            dates.append(col[1].value)
            for p in range(len(names)):
                availability[p].append(process_entry(col[p+2].value))


    print(f'*****')
    # for row in range(1, dataframe1.max_row):
    #     for col in dataframe1.iter_cols(1, dataframe1.max_column):
    #         # print(col[row].value)
    #         pass
    #     # print(f'-----')

    print(f'-----------------------')
    print(dates)
    print(names)
    print(f'---')
    for row in availability:
        print(row)

    return availability, dates, names


if __name__ == "__main__":
    process_notebook("Excel_Files/Test.xlsx")